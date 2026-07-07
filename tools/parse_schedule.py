#!/usr/bin/env python3
"""Parse the USAFL Nationals umpire assignment spreadsheet into app JSON.

Usage: python3 tools/parse_schedule.py "<path to xlsx>" > app/data.js
Re-run whenever a new version of the sheet arrives.
"""
import json
import re
import sys
from openpyxl import load_workbook

XLSX = sys.argv[1] if len(sys.argv) > 1 else "/Users/alison/Desktop/Claude/USAFL Umpire Assignments Prelim Oct 10.xlsx"

# Slot columns 15..34 (1-indexed). Sat 8am-5pm, Sun 8am-1pm, Finals 1pm-4:30pm.
SLOTS = [
    ("sat", "8am"), ("sat", "9am"), ("sat", "10am"), ("sat", "11am"), ("sat", "12pm"),
    ("sat", "1pm"), ("sat", "2pm"), ("sat", "3pm"), ("sat", "4pm"), ("sat", "5pm"),
    ("sun", "8am"), ("sun", "9am"), ("sun", "10am"), ("sun", "11am"), ("sun", "12pm"), ("sun", "1pm"),
    ("finals", "1pm"), ("finals", "2pm"), ("finals", "3pm"), ("finals", "4:30pm"),
]

SECTIONS = [(10, 51, "field"), (52, 72, "goal"), (74, 74, "boundary")]


def classify(raw):
    """Turn a raw cell into a structured slot entry."""
    if raw is None:
        return None
    v = str(raw).strip()
    if not v:
        return None
    tentative = "?" in v
    clean = v.replace("?", "").strip().rstrip("/").strip()
    low = clean.lower()

    if low in ("off", "no games"):
        return {"type": "off", "label": "OFF", "tentative": tentative}
    if low in ("personal", "mvl"):
        return {"type": "personal", "label": "Personal", "tentative": tentative}
    if low == "youth":
        return {"type": "duty", "label": "Youth game", "tentative": tentative}
    if clean.startswith("."):
        return {"type": "coach", "label": clean[1:].strip(), "tentative": tentative}
    if low.startswith("f marshall"):
        return {"type": "duty", "label": "Field Marshall", "tentative": tentative}
    m = re.match(r"^field\s*(\d)$", low)
    if m:
        return {"type": "field", "label": f"Field {m.group(1)}", "field": int(m.group(1)), "tentative": tentative}
    m = re.match(r"^goal\s*(\d)$", low)
    if m:
        return {"type": "goal", "label": f"Field {m.group(1)}", "field": int(m.group(1)), "tentative": tentative}
    m = re.match(r"^bound\s*(\d)$", low)
    if m:
        return {"type": "boundary", "label": f"Field {m.group(1)}", "field": int(m.group(1)), "tentative": tentative}
    m = re.match(r"^ts\s*(\d)$", low)
    if m:
        return {"type": "ts", "label": f"Field {m.group(1)}", "field": int(m.group(1)), "tentative": tentative}
    if re.search(r"\b(pre|play|post|loser|play in)\b", low):
        return {"type": "playing", "label": "With club", "tentative": tentative}
    if "watch" in low:
        return {"type": "watch", "label": "Watching", "tentative": tentative}
    if "coach" in low:
        return {"type": "coach", "label": "Coaching", "tentative": tentative}
    if low.startswith("[") or low in ("tba",):
        return {"type": "maybe", "label": clean.strip("[]"), "tentative": True}
    return {"type": "other", "label": clean, "tentative": tentative}


def main():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb.active
    people = {}
    for start, end, section in SECTIONS:
        for r in range(start, end + 1):
            first = ws.cell(r, 3).value
            last = ws.cell(r, 4).value
            if not first or not last or str(first).strip() == "TBA":
                continue
            first = str(first).strip()
            # "FI LName" column is like "S Arnott" -> drop leading initial
            last_name = re.sub(r"^[A-Z] ", "", str(last).strip())
            key = f"{first.lower()}|{last_name.lower()}"
            club_m = str(ws.cell(r, 5).value or "").strip()
            club_w = str(ws.cell(r, 6).value or "").strip()
            now = str(ws.cell(r, 7).value or "").strip()
            try_ = str(ws.cell(r, 8).value or "").strip()
            commit = str(ws.cell(r, 9).value or "").strip()
            counts = {
                "field": int(ws.cell(r, 10).value or 0),
                "goal": int(ws.cell(r, 11).value or 0),
                "boundary": int(ws.cell(r, 12).value or 0),
                "ts": int(ws.cell(r, 13).value or 0),
                "coach": int(ws.cell(r, 14).value or 0),
            }
            slots = []
            for i, (day, time) in enumerate(SLOTS):
                entry = classify(ws.cell(r, 15 + i).value)
                if entry:
                    entry["day"] = day
                    entry["time"] = time
                    slots.append(entry)

            if key in people:
                p = people[key]
                p["accreditation"][section] = now
                if try_ and try_ != "-":
                    p["try"][section] = try_
                # merge counts: rows repeat totals per section; take max per discipline
                for k in counts:
                    p["counts"][k] = max(p["counts"][k], counts[k])
                # merge slots: prefer concrete assignments over placeholders
                by_key = {(s["day"], s["time"]): s for s in p["slots"]}
                for s in slots:
                    k2 = (s["day"], s["time"])
                    cur = by_key.get(k2)
                    rank = {"field": 5, "goal": 5, "boundary": 5, "ts": 5, "coach": 4, "duty": 4,
                            "playing": 3, "watch": 2, "maybe": 1, "other": 2, "personal": 2, "off": 0}
                    if cur is None or rank.get(s["type"], 0) > rank.get(cur["type"], 0):
                        by_key[k2] = s
                p["slots"] = sorted(by_key.values(), key=lambda s: SLOTS.index((s["day"], s["time"])))
                p["disciplines"].append(section)
            else:
                people[key] = {
                    "id": re.sub(r"[^a-z0-9]+", "-", f"{first} {last_name}".lower()).strip("-"),
                    "firstName": first,
                    "lastName": last_name,
                    "name": f"{first} {last_name}",
                    "club": club_m if club_m and club_m != "N/A" else (club_w if club_w != "N/A" else ""),
                    "isAussie": club_m.upper() == "OZ" or club_w.upper() == "OZ",
                    "accreditation": {section: now},
                    "try": ({section: try_} if try_ and try_ != "-" else {}),
                    "commitment": commit,
                    "counts": counts,
                    "disciplines": [section],
                    "slots": slots,
                }
    roster = sorted(people.values(), key=lambda p: (p["lastName"].lower(), p["firstName"].lower()))
    print("// Auto-generated by tools/parse_schedule.py — do not edit by hand.")
    print("window.UMPIRES = " + json.dumps(roster, indent=1) + ";")
    print("window.SLOT_ORDER = " + json.dumps([{"day": d, "time": t} for d, t in SLOTS]) + ";")


if __name__ == "__main__":
    main()
